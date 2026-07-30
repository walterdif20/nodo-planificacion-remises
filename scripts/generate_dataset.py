import csv
import difflib
import json
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


def clean_text(value):
    return " ".join(str(value or "").strip().split())


def key_text(value):
    text = unicodedata.normalize("NFD", clean_text(value).upper())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    return " ".join(text.split())


def number(value):
    return float(value) if isinstance(value, (int, float)) and math.isfinite(value) else 0.0


def pretty_name(value):
    fixes = {
        "VICUA MACKENNA": "Vicuña Mackenna",
        "GENERAL GEMES": "General Güemes",
        "SAN FCO DEL MONTE DE ORO": "San Francisco del Monte de Oro",
        "JUAN N FERNANDEZ": "Juan N. Fernández",
        "ADOLFO G CHAVES": "Adolfo G. Chaves",
        "GRAL JOSE SAN MARTIN": "General José de San Martín",
        "PTO SAN MARTIN": "Puerto General San Martín",
        "ITAEMB GUAZ": "Itaembé Guazú",
    }
    normalized = key_text(value)
    if normalized in fixes:
        return fixes[normalized]
    words = clean_text(value).lower().title().split()
    particles = {"De", "Del", "La", "Las", "Los", "Y"}
    return " ".join(word.lower() if index and word in particles else word for index, word in enumerate(words))


province_names = {
    "BUENOS AIRES": "Buenos Aires",
    "CATAMARCA": "Catamarca",
    "CHACO": "Chaco",
    "CHUBUT": "Chubut",
    "CORDOBA": "Córdoba",
    "CORRIENTES": "Corrientes",
    "ENTRE RIOS": "Entre Ríos",
    "FORMOSA": "Formosa",
    "JUJUY": "Jujuy",
    "LA PAMPA": "La Pampa",
    "MENDOZA": "Mendoza",
    "MISIONES": "Misiones",
    "NEUQUEN": "Neuquén",
    "RIO NEGRO": "Río Negro",
    "ROSARIO": "Santa Fe",
    "SALTA": "Salta",
    "SAN JUAN": "San Juan",
    "SAN LUIS": "San Luis",
    "SANTA CRUZ": "Santa Cruz",
    "SANTA FE": "Santa Fe",
    "SANTIAGO DEL ESTERO": "Santiago del Estero",
    "TIERRA DEL FUEGO": "Tierra del Fuego",
    "TIERRA DEL FUEGO ANTARTIDA E ISLAS DEL ATLANTICO SUR": "Tierra del Fuego",
    "TUCUMAN": "Tucumán",
}


locality_aliases = {
    "SAN FCO DEL MONTE DE ORO": "SAN FRANCISCO DEL MONTE DE ORO",
    "ADOLFO G CHAVES": "ADOLFO GONZALES CHAVES",
    "CHAVES": "ADOLFO GONZALES CHAVES",
    "GRAL JOSE SAN MARTIN": "GENERAL JOSE DE SAN MARTIN",
    "GENERAL MADARIAGA": "GENERAL JUAN MADARIAGA",
    "GENERAL MANSILLA": "BARTOLOME BAVIO",
    "NUEVA JAMAICA": "VISTA ALEGRE NORTE",
    "PTO SAN MARTIN": "PUERTO GENERAL SAN MARTIN",
}


# Curaduría de nombres operativos que representan barrios, parajes o nombres de uso
# habitual. Los puntos provienen de Georef/IGN o de cartografía abierta; no se usan
# direcciones personales de la hoja Datos.
place_overrides = {
    "BUENOS AIRES|BATAN": (-38.0069, -57.7094, "Buenos Aires", "curated"),
    "BUENOS AIRES|LA DULCE": (-38.283052, -59.202744, "Buenos Aires", "official-alias"),
    "BUENOS AIRES|LAGUNA BRAVA": (-37.861223, -57.981012, "Buenos Aires", "curated"),
    "BUENOS AIRES|LOMA NEGRA": (-36.981005, -60.278602, "Buenos Aires", "official-alias"),
    "BUENOS AIRES|LOS CHANARES": (-33.302086, -66.336855, "San Luis", "curated"),
    "CATAMARCA|SAN ANTONIO LA PAZ": (-28.932482, -65.096917, "Catamarca", "official-alias"),
    "CHUBUT|DIADEMA": (-45.774965, -67.674774, "Chubut", "official-alias"),
    "CORDOBA|ARGUELLO": (-31.346889, -64.253306, "Córdoba", "curated"),
    "CORDOBA|CERRO NORTE": (-31.321253, -64.275297, "Córdoba", "curated"),
    "CORDOBA|CERRITO": (-31.323469, -64.250744, "Córdoba", "curated"),
    "CORDOBA|ITUZAINGO": (-31.465, -64.096389, "Córdoba", "curated"),
    "CORDOBA|VILLA 9 DE JULIO": (-31.327222, -64.2775, "Córdoba", "curated"),
    "CORDOBA|VILLA CORONEL OLMEDO": (-31.481828, -64.137861, "Córdoba", "curated"),
    "MISIONES|ITAEMBE GUAZU": (-27.40569, -55.99271, "Misiones", "curated"),
    "RIO NEGRO|LAGO PELLEGRINI": (-38.706944, -67.973056, "Río Negro", "official"),
    "SAN JUAN|SAN MARTIN": (-31.516685, -68.350119, "San Juan", "official-alias"),
    "SAN JUAN|AMPACAMA": (-31.385645, -68.483708, "San Juan", "curated"),
    "SAN JUAN|MEDIA AGUA": (-31.981447, -68.426673, "San Juan", "official-alias"),
    "SAN LUIS|CUCHI CORRAL": (-33.27167, -66.21993, "San Luis", "curated"),
    "SAN LUIS|HIPOLITO YRIGOYEN": (-33.690184, -65.444874, "San Luis", "curated"),
    "SAN LUIS|VALLE DE LA PANCANTA": (-32.897959, -66.110292, "San Luis", "official-alias"),
    "SANTA FE|ANDINO": (-32.669942, -60.875378, "Santa Fe", "official-alias"),
    "TUCUMAN|SAN RAFAEL": (-26.957603, -65.350551, "Tucumán", "official-alias"),
    "TUCUMAN|TRANCAS": (-26.230908, -65.283447, "Tucumán", "official-alias"),
    "TUCUMAN|BURRUYACU": (-26.500473, -64.743672, "Tucumán", "official-alias"),
}


base_provinces = {
    "PEREZ MILLAN": "Buenos Aires",
    "CORDOBA": "Córdoba",
    "MAR DEL PLATA": "Buenos Aires",
    "SAN MARTIN": "Mendoza",
    "SAN LUIS": "San Luis",
    "ROSARIO": "Santa Fe",
    "OBERA": "Misiones",
    "TILISARAO": "San Luis",
    "CHACABUCO": "Buenos Aires",
    "CENTENARIO": "Neuquén",
    "CAMPANA": "Buenos Aires",
    "GENERAL VILLEGAS": "Buenos Aires",
    "PEDRO LURO": "Buenos Aires",
    "MERCEDES": "Buenos Aires",
    "PERGAMINO": "Buenos Aires",
    "BRANDSEN": "Buenos Aires",
    "MAIPU": "Mendoza",
    "GARUPA": "Misiones",
    "CHIVILCOY": "Buenos Aires",
    "9 DE JULIO": "Buenos Aires",
    "PINAMAR": "Buenos Aires",
    "TRES ARROYOS": "Buenos Aires",
    "MAGDALENA": "Buenos Aires",
    "TORNQUIST": "Buenos Aires",
    "SALADILLO": "Buenos Aires",
    "SAN JUAN": "San Juan",
    "LA PUNTA": "San Luis",
}


def load_geo_catalog(georef_path):
    georef_path = Path(georef_path)
    paths = sorted(georef_path.glob("*.csv")) if georef_path.is_dir() else [georef_path]
    layer_priority = {"localidades": 0, "localidades_censales": 1, "asentamientos": 2}
    records = []
    seen = set()
    for path in paths:
        with path.open("r", encoding="utf-8", newline="") as handle:
            for item in csv.DictReader(handle):
                try:
                    lat = float(item["centroide_lat"])
                    lng = float(item["centroide_lon"])
                except (KeyError, TypeError, ValueError):
                    continue
                if not (-55.2 <= lat <= -21.0 and -74.0 <= lng <= -52.0):
                    continue
                province = province_names.get(
                    key_text(item.get("provincia_nombre")),
                    pretty_name(item.get("provincia_nombre")),
                )
                record = {
                    "province": province,
                    "provinceKey": key_text(province),
                    "primaryName": key_text(item.get("nombre")),
                    "censusName": key_text(item.get("localidad_censal_nombre")),
                    "governmentName": key_text(item.get("gobierno_local_nombre")),
                    "lat": lat,
                    "lng": lng,
                    "layer": path.stem,
                    "layerPriority": layer_priority.get(path.stem, 9),
                    "category": clean_text(item.get("categoria")),
                }
                identity = (
                    record["provinceKey"],
                    record["primaryName"],
                    round(lat, 5),
                    round(lng, 5),
                )
                if record["primaryName"] and identity not in seen:
                    records.append(record)
                    seen.add(identity)
    return records


def build_locator(locations):
    by_province = defaultdict(list)
    primary_global = defaultdict(list)
    census_global = defaultdict(list)
    for location in locations:
        by_province[location["provinceKey"]].append(location)
        primary_global[location["primaryName"]].append(location)
        if location["censusName"]:
            census_global[location["censusName"]].append(location)

    def choose(candidates):
        return min(
            candidates,
            key=lambda item: (
                item["layerPriority"],
                0 if "Localidad" in item["category"] else 1,
                item["primaryName"],
            ),
        )

    def locate(locality, province=None):
        locality_key = locality_aliases.get(key_text(locality), key_text(locality))
        province_key = key_text(province)
        override = place_overrides.get(f"{province_key}|{key_text(locality)}")
        if override:
            lat, lng, corrected_province, source = override
            return {
                "province": corrected_province,
                "provinceKey": key_text(corrected_province),
                "primaryName": locality_key,
                "lat": lat,
                "lng": lng,
                "layer": source,
            }, 0.97 if source.startswith("official") else 0.92, source

        province_pool = by_province.get(province_key, []) if province_key else []
        exact_primary = [item for item in province_pool if item["primaryName"] == locality_key]
        if exact_primary:
            return choose(exact_primary), 1.0, "official"

        exact_census = [item for item in province_pool if item["censusName"] == locality_key]
        if exact_census:
            return choose(exact_census), 0.99, "official-census"

        global_primary = primary_global.get(locality_key, [])
        global_provinces = {item["provinceKey"] for item in global_primary}
        if global_primary and len(global_provinces) == 1:
            return choose(global_primary), 0.97, "official-corrected-province"

        global_census = census_global.get(locality_key, [])
        global_census_provinces = {item["provinceKey"] for item in global_census}
        if global_census and len(global_census_provinces) == 1:
            return choose(global_census), 0.95, "official-census-corrected-province"

        best = None
        best_score = 0.0
        for item in province_pool:
            score = difflib.SequenceMatcher(None, locality_key, item["primaryName"]).ratio()
            if locality_key in item["primaryName"] or item["primaryName"] in locality_key:
                score = max(
                    score,
                    min(len(locality_key), len(item["primaryName"]))
                    / max(len(locality_key), len(item["primaryName"]))
                    + 0.08,
                )
            if score > best_score:
                best = item
                best_score = score
        if best and best_score >= 0.88:
            return best, best_score, "official-fuzzy"
        return None, best_score, "missing"

    return locate


def main(source_path, georef_path, output_path):
    source_path = Path(source_path)
    output_path = Path(output_path)
    locate = build_locator(load_geo_catalog(georef_path))

    book = load_workbook(source_path, data_only=True, read_only=True)
    datos = book["Datos"]
    tabla = book["Tabla"]

    coordinate_pattern = re.compile(r"\((-?\d{1,2}\.\d{4,}),\s*(-?\d{1,3}\.\d{4,})\)")
    source_coordinates = defaultdict(list)
    base_counts = Counter()
    for row in datos.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        base = clean_text(row[18])
        if base:
            base_counts[key_text(base)] += 1
        locality_key = key_text(row[7])
        for value in (row[6], row[19], row[32], row[33]):
            match = coordinate_pattern.search(clean_text(value))
            if locality_key and match:
                source_coordinates[locality_key].append((float(match.group(1)), float(match.group(2))))

    rows = []
    for source_row, row in enumerate(tabla.iter_rows(min_row=2, max_row=355, values_only=True), start=2):
        raw_province = clean_text(row[0])
        raw_locality = clean_text(row[1])
        if not raw_province or not raw_locality or key_text(raw_province).startswith("TOTAL"):
            continue
        source_province = province_names.get(key_text(raw_province), pretty_name(raw_province))
        location, confidence, coordinate_source = locate(raw_locality, source_province)
        if not location and source_coordinates.get(key_text(raw_locality)):
            samples = source_coordinates[key_text(raw_locality)]
            lat = sum(item[0] for item in samples) / len(samples)
            lng = sum(item[1] for item in samples) / len(samples)
            if -55.2 <= lat <= -21.0 and -74.0 <= lng <= -52.0:
                location = {
                    "lat": lat,
                    "lng": lng,
                    "province": source_province,
                    "primaryName": key_text(raw_locality),
                    "layer": "workbook",
                }
                confidence = 0.9
                coordinate_source = "workbook"
        display_province = location.get("province", source_province) if location else source_province
        rows.append({
            "id": f"loc-{source_row}",
            "sourceRow": source_row,
            "province": display_province,
            "sourceProvince": source_province,
            "provinceCorrected": display_province != source_province,
            "locality": pretty_name(raw_locality),
            "urbanTrips": int(number(row[2])),
            "ruralTrips": int(number(row[3])),
            "totalTrips": int(number(row[4])),
            "urbanExcessKm": round(number(row[5])),
            "ruralExcessKm": round(number(row[7])),
            "totalExcessKm": round(number(row[9])),
            "averageExcessKm": round(number(row[10]), 1),
            "lat": round(location["lat"], 6) if location else None,
            "lng": round(location["lng"], 6) if location else None,
            "coordinateConfidence": round(confidence, 3) if location else 0,
            "coordinateSource": coordinate_source,
            "coordinateMatch": pretty_name(location.get("primaryName", raw_locality)) if location else None,
        })

    positive_rows = [row for row in rows if row["totalExcessKm"] > 0]
    max_excess = max(row["totalExcessKm"] for row in positive_rows)
    max_trips = max(row["totalTrips"] for row in positive_rows)
    max_average = max(row["averageExcessKm"] for row in positive_rows)
    for row in positive_rows:
        excess_signal = row["totalExcessKm"] / max_excess
        trip_signal = math.sqrt(row["totalTrips"] / max_trips) if row["totalTrips"] else 0
        average_signal = math.sqrt(row["averageExcessKm"] / max_average) if row["averageExcessKm"] else 0
        row["priorityScore"] = round(100 * (0.65 * excess_signal + 0.25 * trip_signal + 0.10 * average_signal))

    positive_rows.sort(key=lambda item: (item["priorityScore"], item["totalExcessKm"]), reverse=True)
    for rank, row in enumerate(positive_rows, start=1):
        row["rank"] = rank

    existing_bases = []
    for base_key, trips in base_counts.most_common(40):
        province = base_provinces.get(base_key)
        if not province:
            continue
        location, confidence, coordinate_source = locate(base_key, province)
        if not location:
            continue
        existing_bases.append({
            "id": f"base-{len(existing_bases) + 1}",
            "name": pretty_name(base_key),
            "province": location.get("province", province),
            "trips": trips,
            "lat": round(location["lat"], 6),
            "lng": round(location["lng"], 6),
            "coordinateConfidence": round(confidence, 3),
            "coordinateSource": coordinate_source,
        })

    geolocated_count = sum(1 for row in positive_rows if row["lat"] is not None)
    curated_count = sum(1 for row in positive_rows if row["coordinateSource"] in {"curated", "official-alias"})
    corrected_provinces = sum(1 for row in positive_rows if row["provinceCorrected"])
    payload = {
        "meta": {
            "sourceFile": source_path.name,
            "period": "Enero 2026",
            "totalTrips": sum(row["totalTrips"] for row in rows),
            "totalExcessKm": sum(row["totalExcessKm"] for row in rows),
            "analyzedLocalities": len(rows),
            "positiveOpportunityLocalities": len(positive_rows),
            "geolocatedOpportunityLocalities": geolocated_count,
            "coordinateCoveragePercent": round(100 * geolocated_count / len(positive_rows), 1),
            "curatedCoordinateLocalities": curated_count,
            "correctedProvinceLocalities": corrected_provinces,
            "generatedFrom": "Tabla!A2:K355",
            "coordinateSource": "Georef Argentina + curaduría de barrios y equivalencias",
        },
        "methodology": {
            "priority": "65% km excedentes, 25% frecuencia de viajes y 10% excedente promedio, normalizados dentro del período.",
            "simulation": "Ahorro estimado por proximidad: captura configurada × km excedentes × peso lineal según distancia dentro del radio.",
            "caveat": "La simulación es orientativa y no incorpora costos fijos, disponibilidad de prestadores ni restricciones contractuales.",
            "coordinates": "Los nombres se validan primero contra Georef por localidad exacta. Barrios y nombres operativos se resuelven con equivalencias curadas sin exponer direcciones personales.",
        },
        "opportunities": positive_rows,
        "existingBases": existing_bases,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps(payload["meta"], ensure_ascii=False, indent=2))
    missing = [row for row in positive_rows if row["lat"] is None]
    if missing:
        print("Sin coordenadas:")
        for row in missing:
            print(row["sourceRow"], row["locality"], row["province"])


if __name__ == "__main__":
    main(*sys.argv[1:4])
